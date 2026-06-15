#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# Description:
# Processes databases/oncokb/oncokb_precision_oncology_therapies.tsv
#   and reformats it from TSV to JSON importable into MongoDB
# -_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_

import argparse
import json
import re
from pathlib import Path
from typing import List
from openpyxl import load_workbook


class C:
    pass


c = C()
parser = argparse.ArgumentParser(
    description='Processes the OncoKB database of genes related to cancer types and generates a new JSON file importable into MongoDB.')
parser.add_argument(
    '--input', help='XLSX file for the OncoKB precision oncology therapies database', required=True)
parser.add_argument('--output', help='Output JSON file name',
                    required=False, default="oncokb_pot_output.json")

args = parser.parse_args(namespace=c)

valid_genes_file = Path(__file__).with_name('valid_genes_from_hgnc.txt')

with open(valid_genes_file, 'r', encoding='utf-8') as file:
    hgnc_genes = []  # list of valid genes symbols
    hgnc_aliases = {}  # dict of aliases and their valid symbols. If an alias has 2 or more valid symbols it will not appear in dict
    hgnc_previous = {}  # dict of previous and their valid symbols. If an previous has 2 or more valid symbols it will not appear in dict
    aliases_with_multiple_valid_symbols = []
    previous_with_multiple_valid_symbols = []

    content = file.readlines()[1:]
    for line in content:
        if line[-1:] == "\n":
            line = line[:-1]
        values = line.split("\t")
        symbol = values[0].strip().upper()
        if not symbol:
            continue
        hgnc_genes.append(symbol)

        if len(values) > 1:
            aliases = values[1].split(",")
            for alias in aliases:
                alias = alias.strip()
                if not alias:
                    continue
                if alias not in hgnc_aliases:
                    # { "alias" : "symbol"}
                    hgnc_aliases[alias] = symbol
                else:
                    if alias not in aliases_with_multiple_valid_symbols:
                        aliases_with_multiple_valid_symbols.append(alias)
        if len(values) > 2:
            previous = values[2].strip().split(",")
            for prev in previous:
                prev = prev.strip()
                if not prev:
                    continue
                if prev not in hgnc_previous:
                    # { "prev" : "symbol"}
                    hgnc_previous[prev] = symbol
                else:
                    if prev not in previous_with_multiple_valid_symbols:
                        previous_with_multiple_valid_symbols.append(prev)

    for a in aliases_with_multiple_valid_symbols:
        del hgnc_aliases[a]
    for p in previous_with_multiple_valid_symbols:
        del hgnc_previous[p]

valid_genes_set = set(hgnc_genes)

valid_genes_compiled_patern = [
    re.compile(r'\b' + re.escape(gene) + r'\b', flags=re.IGNORECASE)
    for gene in hgnc_genes
]

aliases_compiled_patern = []
for alias in hgnc_aliases:
    try:
        aliases_compiled_patern.append(
            re.compile(r'\b' + re.escape(alias) + r'\b', flags=re.IGNORECASE)
        )
    except Exception:
        pass

previous_compiled_patern = []
for prev in hgnc_previous:
    try:
        previous_compiled_patern.append(
            re.compile(r'\b' + re.escape(prev) + r'\b', flags=re.IGNORECASE)
        )
    except Exception:
        pass


HEADER_ALIASES = {
    "Year of drug’s first FDA-approval a": "fda_first_approval",
    "Year of drug’s first FDA- approval": "fda_first_approval",
    "FDA-approved drug(s) a": "fda_approved_drugs",
    "Precision oncology therapy\nN=86": "precision_oncology_therapy",
    "Precision oncology therapy": "precision_oncology_therapy",
    "FDA-recognized biomarker(s) b": "fda_recognized_biomarkers",
    "FDA drug label listed biomarker(s) b": "fda_recognized_biomarkers",
    "Method of biomarker detection c": "method_of_biomarker_detection",
    "Can a DNA/NGS-based method be used for biomarker detection?": "method_of_biomarker_detection",
    "Can a DNA/NGS-based method be used for biomarker detection? ": "method_of_biomarker_detection",
    "Drug classification  d": "drug_classification",
    "Class of agent(s) c": "drug_classification",
    "Mechanism of action or drug target c": "mechanism_of_action_or_drug_target",
    "Targeted therapy": "targeted_therapy",
}


def normalize_header(value: str) -> str:
    return re.sub(r'\s+', ' ', value.replace("\n", " ")).strip()


def search_gene(biomarker_description: str) -> List[str]:
    genes = set()
    special_cases = {"NTRK1/2/3": ["NTRK1", "NTRK2", "NTRK3"],
                     "BRCA1/2": ["BRCA1", "BRCA2"],
                     "TSC1/2": ["TSC1", "TSC2"],
                     "PDGFRA/B": ["PDGFRA", "PDGFRB"]
                     }
    biomarker_text = biomarker_description.upper()

    for special_case in special_cases:
        if re.search(re.escape(special_case), biomarker_text):
            genes.update(special_cases[special_case])

    for i in range(0, len(hgnc_genes)):
        if re.search(valid_genes_compiled_patern[i], biomarker_text):
            genes.add(hgnc_genes[i])
    a = 0
    for alias in hgnc_aliases:
        if re.search(aliases_compiled_patern[a], biomarker_text):
            genes.add(hgnc_aliases[alias])
        a += 1
    p = 0
    for prev in hgnc_previous:
        if re.search(previous_compiled_patern[p], biomarker_text):
            genes.add(hgnc_previous[prev])
        p += 1

    # Guardrail: only return official HGNC approved symbols.
    genes = [gene for gene in genes if gene in valid_genes_set]
    genes.sort()
    return genes


if __name__ == '__main__':
    input_file = c.input  # type: ignore
    output_file = c.output  # type: ignore
    json_content = []
    pattern = r"\?\s*\d+\s*\*"

    print("Processing file...")
    workbook = load_workbook(filename=input_file, data_only=True)
    worksheet = workbook["FDA-Approved Oncology Therapies"] if "FDA-Approved Oncology Therapies" in workbook.sheetnames else workbook.active
    if worksheet is None:
        raise ValueError("Could not get a valid worksheet from the XLSX file")

    rows = worksheet.iter_rows(values_only=True)
    raw_headers = next(rows)
    if raw_headers is None:
        raise ValueError("The XLSX file has no headers")

    selected_columns = []
    for idx, raw_header in enumerate(raw_headers):
        if raw_header is None:
            continue
        normalized = normalize_header(str(raw_header))
        mapped = HEADER_ALIASES.get(normalized)
        if mapped:
            selected_columns.append((idx, mapped))

    if not selected_columns:
        raise ValueError("No valid columns were recognized in the XLSX file")

    for row_record in rows:
        if row_record is None:
            continue
        if all(value is None for value in row_record):
            continue

        json_record = {}
        for idx, header_name in selected_columns:
            value = row_record[idx] if idx < len(row_record) else ""
            normalized_value = "" if value is None else str(value)
            normalized_value = normalized_value.replace("\n", " ")

            if header_name == "drug_classification" and normalized_value.upper() == "NA":
                normalized_value = ""
            if header_name == "fda_first_approval":
                if re.search(pattern, normalized_value):
                    normalized_value = normalized_value.replace("?", "-")
                normalized_value = normalized_value.replace("*", "")

            json_record[header_name] = normalized_value

            if header_name == "fda_recognized_biomarkers":
                json_record["hgnc_symbol"] = search_gene(normalized_value)

        json_content.append(json_record)

    with open(output_file, "w", encoding="utf-8") as jsonfile:
        json.dump(json_content, jsonfile, ensure_ascii=False)

    workbook.close()
    print("Done!")
